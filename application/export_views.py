# Copyright (C) 2026 Redcar & Cleveland Borough Council
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, version 3.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""
Server-side export views for the Members Enquiries System.

These views handle exporting filtered enquiry data in various formats,
supporting the full dataset regardless of pagination.
"""

import csv
import io
import json
from datetime import datetime
from urllib.parse import urlencode

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_http_methods

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .class_views import EnquiryFilterMixin
from .forms import EnquiryFilterForm
from .models import Enquiry
from .services import EnquiryFilterService
from .date_range_service import DateRangeService
from .search_service import EnquirySearchService


class ExportDataProcessor:
    """Process enquiry data for export in various formats."""

    def __init__(self, request_data, user):
        self.request_data = request_data
        self.user = user
        self.filter_mixin = EnquiryFilterMixin()

    NOT_ASSIGNED = "Not assigned"

    # Field name to ORM lookup mapping for simple filters
    _FIELD_LOOKUPS = {
        "member": "member",
        "admin": "admin",
        "section": "section",
        "job_type": "job_type",
        "contact": "contact",
        "ward": "member__ward",
    }

    def get_filtered_queryset(self):
        """Get the complete filtered queryset (not paginated)."""
        filter_form = EnquiryFilterForm(self.request_data)
        filter_form.is_valid()

        queryset = Enquiry.objects.select_related(
            "member", "admin__user", "section", "job_type", "contact"
        ).order_by("-created_at")

        queryset = DateRangeService.apply_date_filters(queryset, filter_form)

        if filter_form.is_valid():
            queryset = self._apply_status_filter(queryset, filter_form)
            queryset = self._apply_field_filters(queryset, filter_form)

        search_value = self.request_data.get("search[value]", "").strip()
        if search_value:
            queryset = self._apply_search(queryset, search_value)

        return queryset, filter_form

    @staticmethod
    def _apply_status_filter(queryset, filter_form):
        """Apply status filter to queryset."""
        status = filter_form.cleaned_data.get("status")
        if not status:
            return queryset
        if status == "open":
            return queryset.filter(status__in=["new", "open"])
        return queryset.filter(status=status)

    @classmethod
    def _apply_field_filters(cls, queryset, filter_form):
        """Apply simple field filters from the form."""
        for field_name, lookup in cls._FIELD_LOOKUPS.items():
            value = filter_form.cleaned_data.get(field_name)
            if value:
                queryset = queryset.filter(**{lookup: value})
        return queryset

    def _apply_search(self, queryset, search_value):
        """Apply search to queryset using centralized search service."""
        return EnquirySearchService.apply_search(queryset, search_value)

    @staticmethod
    def _extract_date(value):
        """Extract date from a datetime or return as-is if already a date."""
        return value.date() if hasattr(value, "date") else value

    @staticmethod
    def _format_date(value):
        """Format a date as DD/MM/YYYY string."""
        return value.strftime("%d/%m/%Y")

    @staticmethod
    def _calculate_overdue_days(due_date, today, status):
        """Calculate overdue business days for open enquiries."""
        if status == "closed":
            return "-"
        if due_date >= today:
            return "-"
        from application.utils import calculate_business_days

        overdue_business_days = calculate_business_days(due_date, today)
        if overdue_business_days and overdue_business_days > 0:
            return str(overdue_business_days)
        return "-"

    @staticmethod
    def _calculate_resolution_time(enquiry):
        """Calculate resolution time in business days for closed enquiries."""
        if enquiry.status != "closed" or not enquiry.closed_at:
            return "-"
        from application.utils import calculate_business_days

        business_days = calculate_business_days(enquiry.created_at, enquiry.closed_at)
        if business_days is not None:
            return str(business_days)
        return "-"

    @staticmethod
    def _get_admin_display(enquiry):
        """Get the admin display name for an enquiry."""
        if not enquiry.admin or not enquiry.admin.user:
            return "Unassigned"
        return enquiry.admin.user.get_full_name() or enquiry.admin.user.username

    @classmethod
    def _build_enquiry_row(cls, enquiry, today):
        """Build a single export row from an enquiry."""
        created_date = cls._extract_date(enquiry.created_at)
        updated_date = cls._extract_date(enquiry.updated_at)
        due_date = cls._extract_date(enquiry.due_date)

        closed_date = None
        if enquiry.status == "closed" and enquiry.closed_at:
            closed_date = cls._extract_date(enquiry.closed_at)

        return {
            "reference": enquiry.reference or "No Ref",
            "title": enquiry.title,
            "member": enquiry.member.full_name,
            "section": enquiry.section.name if enquiry.section else cls.NOT_ASSIGNED,
            "job_type": enquiry.job_type.name if enquiry.job_type else cls.NOT_ASSIGNED,
            "contact": enquiry.contact.name if enquiry.contact else cls.NOT_ASSIGNED,
            "status": (
                "Open" if enquiry.status == "new" else enquiry.get_status_display()
            ),
            "admin": cls._get_admin_display(enquiry),
            "created": cls._format_date(created_date),
            "updated": cls._format_date(updated_date),
            "due_date": cls._format_date(due_date),
            "overdue_days": cls._calculate_overdue_days(
                due_date, today, enquiry.status
            ),
            "closed": cls._format_date(closed_date) if closed_date else "-",
            "resolution_time": cls._calculate_resolution_time(enquiry),
        }

    def get_export_data(self):
        """Get data formatted for export."""
        queryset, filter_form = self.get_filtered_queryset()
        dynamic_title = EnquiryFilterService.generate_dynamic_title(filter_form)
        today = timezone.now().date()

        export_data = [self._build_enquiry_row(enquiry, today) for enquiry in queryset]

        return export_data, dynamic_title, queryset.count()


@login_required
@require_http_methods(["GET"])
def export_enquiries_csv(request):
    """Export filtered enquiries to CSV format."""
    processor = ExportDataProcessor(request.GET, request.user)
    export_data, dynamic_title, _ = processor.get_export_data()

    # Create CSV response
    response = HttpResponse(content_type="text/csv")
    filename = f"{dynamic_title}_Export_{datetime.now().strftime('%Y-%m-%d')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Write CSV data
    writer = csv.DictWriter(
        response,
        fieldnames=[
            "reference",
            "title",
            "member",
            "section",
            "job_type",
            "contact",
            "status",
            "admin",
            "created",
            "updated",
            "due_date",
            "overdue_days",
            "closed",
            "resolution_time",
        ],
    )

    writer.writeheader()
    writer.writerows(export_data)

    return response


@login_required
@require_http_methods(["GET"])
def export_enquiries_excel(request):
    """Export filtered enquiries to Excel format."""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse(
            {"error": "Excel export not available - openpyxl not installed"}, status=500
        )

    processor = ExportDataProcessor(request.GET, request.user)
    export_data, dynamic_title, _ = processor.get_export_data()

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiries Export"

    # Define headers
    headers = [
        "Reference",
        "Title",
        "Member",
        "Section",
        "Job Type",
        "Contact",
        "Status",
        "Admin",
        "Created",
        "Updated",
        "Due Date",
        "Overdue Days",
        "Closed",
        "Resolution Time",
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data
    field_mapping = [
        "reference",
        "title",
        "member",
        "section",
        "job_type",
        "contact",
        "status",
        "admin",
        "created",
        "updated",
        "due_date",
        "overdue_days",
        "closed",
        "resolution_time",
    ]

    for row, enquiry_data in enumerate(export_data, 2):
        for col, field in enumerate(field_mapping, 1):
            ws.cell(row=row, column=col, value=enquiry_data[field])

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15

    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"{dynamic_title}_Export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
@require_http_methods(["GET"])
def get_export_info(request):
    """Get information about the current export (count, title) via AJAX."""
    processor = ExportDataProcessor(request.GET, request.user)
    _, dynamic_title, total_count = processor.get_export_data()

    return JsonResponse(
        {
            "title": dynamic_title,
            "count": total_count,
            "csv_url": f"/api/export/csv/?{urlencode(request.GET)}",
            "excel_url": f"/api/export/excel/?{urlencode(request.GET)}",
        }
    )
